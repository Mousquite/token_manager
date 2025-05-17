# src/excel_manager.py

import openpyxl
import os
from datetime import datetime
import pandas as pd


class ExcelManager:
    def __init__(self, filepath="tokens.xlsx"):
        self.filepath = filepath
        self.workbook = None
        self.sheet = None
        self.headers = []
        self.dirty = False
        self.df = None

    def load_excel(self):
        print("Chargement du fichier...")
        if not os.path.exists(self.filepath):
            print(f"Fichier {self.filepath} non trouvé, création...")
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Tokens"
            sheet.append(["contract_address", "token_id"])  # clé minimale
            workbook.save(self.filepath)

        self.workbook = openpyxl.load_workbook(self.filepath)
        self.sheet = self.workbook.active
        #self.df = pd.read_excel(self.filepath) equivalent panda non utilisé

        # Normalisation des en-têtes
        self.headers = [
            str(cell.value).strip().lower() for cell in self.sheet[1] if cell.value is not None
        ]

        # Création de self.df à partir de openpyxl
        data = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(self.headers, row)))
        self.df = pd.DataFrame(data)

        print(f"Fichier chargé avec {len(self.df)} lignes.")

    def get_all_tokens(self):
        tokens = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            token = {self.headers[i]: row[i] for i in range(len(self.headers))}
            tokens.append(token)
        return tokens

    def update_token_field(self, row_idx, field, value):
        field = field.strip().lower()
        if field not in self.headers:
            raise ValueError(f"Champ {field} introuvable.")
        col_idx = self.headers.index(field) + 1
        self.sheet.cell(row=row_idx + 2, column=col_idx, value=value)
        self.dirty = True

    def update_last_scraped(self, row_idx):
        if "last_scraped" not in self.headers:
            raise ValueError("Champ 'last_scraped' non trouvé.")
        col_idx = self.headers.index("last_scraped") + 1
        today = datetime.now().strftime("%Y-%m-%d")
        self.sheet.cell(row=row_idx + 2, column=col_idx, value=today)
        self.dirty = True

    def update_from_table(self, table_widget):
        self.headers = [table_widget.horizontalHeaderItem(col).text() for col in range(table_widget.columnCount())]
    
        # Efface la feuille existante
        self.sheet.delete_cols(1, self.sheet.max_column)
        self.sheet.delete_rows(1, self.sheet.max_row)

        # Réécrit les en-têtes
        self.sheet.append(self.headers)


        # Réécrit les données ligne par ligne
        for row in range(table_widget.rowCount()):
            for col in range(table_widget.columnCount()):
                item = table_widget.item(row, col)
                value = item.text() if item else ""
                self.sheet.cell(row=row+2, column=col+1, value=value)
        self.dirty = True

    def save_excel(self):
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        wb = Workbook()
        ws = wb.active

        for r in dataframe_to_rows(self.df, index=False, header=True):
            ws.append(r)

        wb.save(self.filepath)
        print(f"Modifications sauvegardées dans {self.filepath}.")
        self.dirty = False

    def is_dirty(self):
        return self.dirty
    
    def get_all_data(self):
        if self.df is None:
            raise ValueError("Aucune donnée en mémoire.")
        return self.df.fillna("").to_dict(orient="records")

    def import_table(self, tnew: pd.DataFrame):
        print(">>> Colonnes de newtokens.xlsx :", list(tnew.columns))

        if self.df is None:
            raise ValueError("Aucune table de référence chargée.")

        try:
            def clean_keys(df):
                df = df.copy()

                def extract_from_url(url):
                    try:
                        parts = url.split('/')
                        chain = parts[4]
                        contract_address = parts[5]
                        token_id = str(int(float(parts[6])))  # standardisation du token_id
                        return chain, contract_address, token_id
                    except Exception:
                        return None, None, None

                for idx, row in df.iterrows():
                    url = row.get("url")
                    chain, contract_address, token_id = extract_from_url(str(url)) if pd.notna(url) else (None, None, None)
                    if contract_address:
                        df.at[idx, "contract_address"] = contract_address
                    if token_id:
                        df.at[idx, "token_id"] = token_id
                    if chain:
                        df.at[idx, "chain"] = chain

                return df

            def get_key(df):
                return df["contract_address"].astype(str) + "_" + df["token_id"].astype(str)

            # Nettoyage et normalisation
            t1 = clean_keys(self.df)
            tnew = clean_keys(tnew)

            for col in ["✔", "checked", "Unnamed: 0", "Unnamed: 1", "Unnamed: 2",]:
                if col in tnew.columns:
                    tnew = tnew.drop(columns=[col])
                if col in self.df.columns:
                    self.df = self.df.drop(columns=[col])
            for col in ["✔"]:
                if col in t1.columns:
                    t1 = t1.drop(columns=[col])

            # Supprimer toute colonne "checked" de tnew (ne doit pas être importée)
            if "checked" in tnew.columns:
                tnew = tnew.drop(columns=["checked"])

            # Clés d'identification
            t1_keys = get_key(t1)
            tnew_keys = get_key(tnew)

            # Harmonise les colonnes entre t1 et tnew
            for col in tnew.columns:
                if col not in t1.columns:
                    t1[col] = None
            for col in t1.columns:
                if col not in tnew.columns:
                    tnew[col] = None

            # Index rapide pour mise à jour
            t1_index = {key: i for i, key in enumerate(t1_keys)}
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            for idx, row in tnew.iterrows():
                key = f"{row['contract_address']}_{row['token_id']}"

                if key in t1_index:
                    # Mise à jour : champs non vides + non verrouillés
                    i = t1_index[key]
                    for col in tnew.columns:
                        if pd.notna(row[col]) and row[col] != "":
                            try:
                                col_index = self.df.columns.get_loc(col)
                            except KeyError:
                                continue  # colonne introuvable (sécurité)

                            # Vérification du verrou (si la table a accès à locked_cells via self.table)
                            if hasattr(self, "table") and hasattr(self.table, "locked_cells"):
                                if (i, col_index) in self.table.locked_cells:
                                    continue
                            t1.at[i, col] = row[col]

                    t1.at[i, "last_scrape_date"] = now
                else:
                    # Nouveau token → ajout
                    new_row = {col: row.get(col) if pd.notna(row.get(col)) else None for col in t1.columns}
                    new_row["last_scrape_date"] = now
                    new_row_df = pd.DataFrame([new_row], columns=t1.columns)
                    t1 = pd.concat([t1, new_row_df], ignore_index=True)

            self.df = t1
            print(">>> fin de l'importation succès")
            print(">>> Colonnes actuelles après import :", list(self.df.columns))


        except Exception as e:
            print(">>> Erreur dans import_table :", e)